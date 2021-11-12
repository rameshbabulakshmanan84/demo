import shutil
from openpyxl import workbook
from openpyxl.reader.excel import load_workbook
from datetime import date
from openpyxl.drawing.image import Image
import os
from num2words import num2words
os.chdir('C:/Users/firer/Documents/Attra/Internet allowance/Oct')

summary_wb = load_workbook(filename='Resourceroosterdetails.xlsx')
summary_ws = summary_wb['Sheet1']

summary_wb1 = load_workbook(filename='Internetallowance_Oct.xlsx', )
summary_ws1 = summary_wb1['Sheet1']

for i in range(2, summary_ws.max_row + 1):
    emp_id = str(summary_ws.cell(row=i, column=1).value)
    # print("Rooster",emp_id)
    for j in range(2, summary_ws1.max_row + 1):
        empid_1 = str(summary_ws1.cell(row=j, column=1).value)      
        print(emp_id,empid_1)
        print("allowance",empid_1)
        if empid_1 == emp_id:
            print(emp_id, empid_1)
            #
            # mgr =summary_ws.cell(row=i, column=3).value
            # summary_ws1[j5]=mgr
            summary_ws1.cell(j, 5, value=summary_ws.cell(row=i, column=3).value)
            # summary_wb1.save()
            # print(mgr)
            # Program =summary_ws.cell(row=i,column=4).value
            summary_ws1.cell(j, 6, value=summary_ws.cell(row=i, column=4).value)
            summary_wb1.save('Internetallowance_Oct_v01_report.xlsx')
            # print(Program)
            break
