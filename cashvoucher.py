import shutil
from openpyxl import workbook
from openpyxl.reader.excel import load_workbook
from datetime import date
from openpyxl.drawing.image import Image
import os
from num2words import num2words
os.chdir('C:/Users/firer/Documents/Attra/Internet allowance/Oct')

def update_voucher(file_name1, emp_name1, emp_id1, amount1):
    shutil.copyfile('template.xlsx', file_name1)
    wb = load_workbook(filename=file_name1)
    ws = wb['Voucher']
    ws.cell(9, 3, value=emp_name1)
    ws.cell(9, 6, value=emp_id1)
    ws.cell(14, 6, value=amount1)
    amount_words = (num2words(amount, lang='en')) + ' only'
    amount_words = amount_words.capitalize()
    ws.cell(23, 4, value=amount_words)
    ws.cell(29, 2, value=emp_name)
    ws.cell(29, 4, value='Vivek Venugopal')
    img = Image('attra.png')
    ws.add_image(img, 'B2')
    wb.save(file_name1)
    wb.close()


summary_wb = load_workbook(filename='Internetallowance_Oct.xlsx')
summary_ws = summary_wb['Sheet1']

for i in range(2, summary_ws.max_row + 1):
    emp_id = str(summary_ws.cell(row=i, column=1).value)
    emp_name = str(summary_ws.cell(row=i, column=2).value)
    amount = summary_ws.cell(row=i, column=3).value
    print(amount)
    file_name = emp_id + '_' + emp_name.replace(" ", "") + '_cashvoucher.xlsx'
    # print(file_name)
    update_voucher(file_name, emp_name, emp_id, amount)
